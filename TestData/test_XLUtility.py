import openpyxl

class TestXLUtility:

    def __init__(self, driver):
        self.driver = driver

    def getRowCount(file, sheetname):
        workbook = openpyxl.load_workbook(file)
        #sheet = workbook.get_sheet_by_name(sheetname)
        sheet = workbook[sheetname]
        return(sheet.max_row)

    def getColumnCount(file, sheetname):
        workbook = openpyxl.load_workbook(file)
        #sheet = workbook.get_sheet_by_name(sheetname)
        sheet = workbook[sheetname]
        return (sheet.max_column)

    def readData(file,sheetname,rownum,columno):
        workbook = openpyxl.load_workbook(file)
        #sheet = workbook.get_sheet_by_name(sheetname)
        sheet = workbook[sheetname]
        return(sheet.cell(row=rownum, column=columno).value)

    def writeData(file,sheetname,rownum,columno,data):
        workbook = openpyxl.load_workbook(file)
        #sheet = workbook.get_sheet_by_name(sheetname)
        sheet = workbook[sheetname]
        sheet.cell(row=rownum, column=columno).value = data
        workbook.save(file)
